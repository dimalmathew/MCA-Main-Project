
import json
import os

import openpyxl
from django.db import connection,DatabaseError
from django.db.models import Max, Sum
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, Http404
from django.shortcuts import render, redirect

from ultimatix.utils import render_to_pdf
from .models import *
from datetime import datetime
import calendar
#import datetime
from ultimatix import config
import numpy as np
from django.db.models import Q
# Create your views here.
def dictfetchall(cursor):
    """Return all rows from a cursor as a dict"""
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

def home_view(request):
    cursor = connection.cursor()
    try:
        logged_in = request.session['eid']
    except:
        logged_in = False

    if logged_in:
        date=datetime.now()
        #date=datetime.strptime('2020-01-01','%Y-%m-%d')
        if (date.day == calendar.monthrange(date.year, date.month)[1]):
            ob=Employee.objects.all().exclude(e_id='1022890')
            ed=str(calendar.monthrange(date.year, date.month)[1])
            yr=str(date.year)
            mn=str(date.month)
            if len(mn)==1:
                mn='0'+mn
            sdate=yr+"-"+mn+"-01"
            #print(sdate+" "+edate)
            sm=yr+"-"+mn
            #em=yr+"-"+str(date.month+1)
            yr2=str(date.year)
            #print('date.month'+str(date.month))
            if date.month==12:
                mn2='01'
                yr2=str(date.year+1)
            else:
                mn2=str(date.month+1)
                if len(mn2)==1:
                    mn2='0'+mn2
            em=yr2+"-"+mn2
            edate = yr + "-" + mn + "-" + ed
            print(sdate+" "+edate)
            print(sm + " "+em)

            busdays=np.busday_count(sm,em)
            print(busdays)

            for o in ob:
                #print(o.e_id)

                sql='''
                select count(*) from ultimatix_timesheet where eid_id=%s and tdate >=%s
                and tdate<=%s and tstate='P'
                '''
                cursor.execute(sql,[o.e_id,sdate,edate])
                res=cursor.fetchone()
                p=int(res[0])

                sql='''
                select count(*) from ultimatix_timesheet where eid_id=%s and tdate >=%s
                and tdate<=%s and tstate='H'                
                '''
                cursor.execute(sql,[o.e_id,sdate,edate])
                res=cursor.fetchone()
                h=0.5*int(res[0])

                sql='''
                select count(*) from ultimatix_timesheet where eid_id=%s and tdate >=%s
                and tdate<=%s and tstate='A'                
                '''
                cursor.execute(sql,[o.e_id,sdate,edate])
                res=cursor.fetchone()
                a=int(res[0])

                sql='''
                select sum(nof) as sum from ultimatix_leave where eid_id=%s and sdate>=%s and edate<=%s
                and status='C'            
                '''
                cursor.execute(sql,[o.e_id,sdate,edate])
                #cursor.execute(sql)
                res=cursor.fetchone()
                #print(str(res))
                #print(int(res[0]))
                l=0
                if res[0]:
                    #print('record for : '+str(o.e_id))
                    l=int(res[0])
                #if str(o.e_id)=='1022892':
                tot=(p+h)+(a-l)
                print(str(p)+" "+str(h)+" "+str(a)+" "+str(l)+" "+str(tot))
                ctc= float(Employee_sal.objects.get(es_eid_id=o.e_id,es_status='Y').es_ctc)
                ctc_pm=float(ctc/12)
                ctc_pd=float(ctc_pm/int(calendar.monthrange(date.year, date.month)[1]))
                lop=0
                if tot<busdays:
                    lop=(float(busdays)-float(tot))*ctc_pd
                bs = ctc_pm * .40
                conv = ctc_pm * .18
                hra = ctc_pm * .12
                city = ctc_pm * .10
                sundry = ctc_pm * .06
                ptax = ctc_pm * .04
                pf = ctc_pm * .02
                esis = ctc_pm * .08
                r=Wage.objects.filter(eid_id=o.e_id,sdate=sdate,edate=edate,status='Y')
                if not r:
                    Wage.objects.create(eid_id=o.e_id,sdate=sdate,edate=edate,ctc=ctc,ctc_pm=round(ctc_pm,2),
                                        bs=round(float(bs),2),conv=round(float(conv),2),hra=round(float(hra),2),city=round(float(city),2),
                                        sundry=round(float(sundry),2),ptax=round(float(ptax),2),pf=round(float(pf),2),
                                        esis=round(float(esis),2),lop=round(float(lop),2),status='Y')
        elif date.day==1 and date.month in(1,4,7,10):
            ob = Employee.objects.all().exclude(e_id='1022890')
            if date.month == 1:
                for o in ob:
                    o.e_cl=2.0
                    o.e_sl=2.0
                    o.e_el=4.0
                    o.e_fl=2.0
                    o.save()
                    print('flexi and other leave credit')
            else:
                for o in ob:
                    r=Employee.objects.filter(e_id=str(o.e_id),e_cl=2.0,e_sl=2.0)
                    if not r:
                        o.e_cl=2.0
                        o.e_sl=2.0
                        o.e_el=float(o.e_el)+4.0
                        o.save()
                        print('credit other leave')
        else:
            print('not month end')
        news = News.objects.all().order_by('-nid')
        user = Employee.objects.get(e_id=logged_in)
        sql = '''
        select ud.d_role
        from ultimatix_employee ue
        left outer join ultimatix_employee_desig ued on ued.ed_eid_id = ue.e_id
        left outer join ultimatix_designation ud on ud.d_id = ued.ed_did_id
        where ued.ed_status = 'Y' and ue.e_id = %s
        '''
        cursor.execute(sql, [logged_in])
        res = cursor.fetchone()
        if str(res[0])=='0':
            role=0
        elif str(res[0])=='2':
            role=2
        elif str(res[0])=='1':
            role=1
        cursor.close()
        if role==0:
            return render(request, 'ultimatix/admin/adminhome.html',{'user':user,'news':news})
        else:
            return render(request,'ultimatix/employee/emphome.html',{'user':user,'news':news,'role':role})
    else:
        request.session['eid'] = None
        cursor.close()
        return redirect('ultimatix:show_login')





def designation_view(request):
    if request.is_ajax() and request.GET:
        try:
            cd = request.GET.get('cd')
            #print("code : "+cd)
            cursor = connection.cursor()
            sql = """
            select d_cd,d_desc,d_min,d_max,d_role from ultimatix_designation 
        	where d_cd=%s
            """
            cursor.execute(sql,[cd])
            #print("cursor : "+str(cursor))
            dict = dictfetchall(cursor)
            #print(str(dict))
            r = json.dumps(dict[0])
            cursor.close()
            return JsonResponse(r, safe=False)
        except DatabaseError  as e:
            print('Database error : ' + str(e))
        finally:
            if cursor:
                cursor.close()


    elif request.method=='POST':
        # return HttpResponse('submitted')
        cd = request.POST.get("cd")
        desc = request.POST.get("desc")

        if (request.POST.get("minexp").strip()) == '':
            min_exp = None
        else:
            min_exp = int(request.POST.get("minexp"))

        if (request.POST.get("maxexp").strip()) == '':
            max_exp = ''
        else:
            max_exp = int(request.POST.get("maxexp"))

        r_type = request.POST.get("role_select")
        # print('r_type : '+'*'+str(r_type)+'*')
        r_num = -1
        if r_type == "Employee":
            r_num = 1
        elif r_type == "HR":
            r_num = 2
        if request.POST.get("add"):
            ob=Designation.objects.filter(d_cd=cd)
            if ob:
                logged_in = request.session['eid']
                user = Employee.objects.get(e_id=logged_in)
                obj = Designation.objects.all().exclude(d_role=0)
                return render(request, 'ultimatix/admin/designation.html',
                             {'error': ' Designation already exists !','obj':obj,'user':user})
            else:
                try:
                    print('Entered for new save')
                    cursor = connection.cursor()
                    sql="""
                    insert into ultimatix_designation(d_cd,d_desc,d_min,d_max,d_role)
                    values(%s,%s,%s,%s,%s)
                    """
                    cursor.execute(sql, (cd, desc, min_exp, max_exp, r_num))
                    connection.commit()
                    cursor.close()
                    #obj=Designation.objects.create(d_cd=cd,d_desc=desc,d_min=min_exp,d_max=max_exp,d_role=r_num)
                    #obj.save()
                    logged_in = request.session['eid']
                    user = Employee.objects.get(e_id=logged_in)
                    obj = Designation.objects.all().exclude(d_role=0)
                    return render(request, 'ultimatix/admin/designation.html',
                              {'success': ' Data saved successfully!','obj':obj,'user':user})
                except DatabaseError  as e:
                    print('Database error : '+str(e))
                    connection.rollback()
                    obj = Designation.objects.all().exclude(d_role=0)
                    logged_in = request.session['eid']
                    user = Employee.objects.get(e_id=logged_in)
                    return render(request, 'ultimatix/admin/designation.html',
                              {'error': ' Database error.Please contact system admin!','obj':obj,'user':user})
                finally:
                    if cursor:
                        cursor.close()
        elif request.POST.get("update"):
            ob = Designation.objects.filter(d_cd=cd)
            if ob:
                logged_in = request.session['eid']
                user = Employee.objects.get(e_id=logged_in)
                Designation.objects.filter(d_cd=cd).update(d_desc=desc,d_min=min_exp,d_max=max_exp,d_role=r_num)
                obj = Designation.objects.all().exclude(d_role=0)
                return render(request, 'ultimatix/admin/designation.html',
                             {'success': ' Data saved successfully!','obj':obj,'user':user})
            else:
                logged_in = request.session['eid']
                user = Employee.objects.get(e_id=logged_in)
                obj = Designation.objects.all().exclude(d_role=0)
                return render(request, 'ultimatix/admin/designation.html',
                             {'error': ' Designation does not exist !','obj':obj,'user':user})
    else:
        logged_in = request.session['eid']
        user = Employee.objects.get(e_id=logged_in)
        obj=Designation.objects.all().exclude(d_role=0)
        return render(request,'ultimatix/admin/designation.html',{'obj':obj,'user':user})



def roles_view(request):
    if request.is_ajax() and request.GET:
        cd=request.GET.get('cd')
        try:
            cursor = connection.cursor()
            sql="""
            select r_cd,r_desc,r_active from ultimatix_project_roles
            where r_cd=%s
            """
            cursor.execute(sql, [cd])
            dict = dictfetchall(cursor)
            r = json.dumps(dict[0])
            cursor.close()
            return JsonResponse(r, safe=False)
        except DatabaseError  as e:
            print('Database error : ' + str(e))
        finally:
            if cursor:
                cursor.close()

    elif request.method=='POST':
        cd=request.POST.get("role_cd")
        desc=request.POST.get("role_desc")
        active=str(request.POST.get("chk_act"))
        updt=str(datetime.now())
        if active !='Y':
            active='N'
        if request.POST.get("add"):
            p_obj=Project_roles.objects.filter(r_cd=cd)
            if p_obj:
                obj = Project_roles.objects.all()
                logged_in = request.session['eid']
                user = Employee.objects.get(e_id=logged_in)
                return render(request, 'ultimatix/admin/roles.html',
                              {'error': ' Project role already exists !', 'obj': obj,'user':user})
            else:
                ob=Project_roles.objects.create(r_cd=cd,r_desc=desc,r_active=active,r_updt=updt)
                ob.save()

                obj = Project_roles.objects.all()
                logged_in = request.session['eid']
                user = Employee.objects.get(e_id=logged_in)
                return render(request, 'ultimatix/admin/roles.html', {'success': ' Data saved successfully!','obj': obj,'user':user})
        elif request.POST.get("update"):
            ob = Project_roles.objects.filter(r_cd=cd)
            if ob:
                logged_in = request.session['eid']
                user = Employee.objects.get(e_id=logged_in)
                Project_roles.objects.filter(r_cd=cd).update(r_desc=desc,r_active=active,r_updt=updt)
                obj = Project_roles.objects.all()
                return render(request, 'ultimatix/admin/roles.html',
                             {'success': ' Data saved successfully!','obj':obj,'user':user})
            else:
                logged_in = request.session['eid']
                user = Employee.objects.get(e_id=logged_in)
                obj = Project_roles.objects.all()
                return render(request, 'ultimatix/admin/roles.html',
                             {'error': ' Role does not exist !','obj':obj,'user':user})

    else:
        logged_in = request.session['eid']
        user = Employee.objects.get(e_id=logged_in)
        obj = Project_roles.objects.all()
        return render(request,'ultimatix/admin/roles.html',{'obj':obj,'user':user})


def regemp_view(request):
    if request.method == 'POST':
        if(request.POST.get('register')):
            #return HttpResponse('Hi')
            fname=request.POST.get('fname')
            lname=request.POST.get('lname')
            sname=request.POST.get('surname')
            dob=request.POST.get('dob')
            bgroup=request.POST.get('bgroup')
            mstatus=request.POST.get('mstatus')
            nationality=request.POST.get('nationality')
            disb=request.POST.get('chk_disb')
            if not disb:
                disb='N'

            gender=request.POST.get('gender')

            paddr1=request.POST.get('paddr1')
            paddr2=request.POST.get('paddr2')
            paddr3=request.POST.get('paddr3')
            caddr1=request.POST.get('caddr1')
            caddr2=request.POST.get('caddr2')
            caddr3=request.POST.get('caddr3')
            email=request.POST.get('email')
            mnumber=request.POST.get('mnumber')

            doj=request.POST.get('doj')
            qfn=request.POST.get('qfn')
            prev_exp=str(request.POST.get('prev_exp'))
            if not prev_exp:
                prev_exp=0
            prev_cmp=request.POST.get('prev_cmp')


            idob = id_select.objects.get(id_desc='e_id')
            id = idob.id_val

            status='Y'
            pwd=str(id)+'@'+dob
            des=request.POST.get('desig_select')
            sal=request.POST.get('salary')
            des_id=Designation.objects.get(d_cd=des).d_id
            des_sdate=str(datetime.today().strftime('%Y-%m-%d'))
            #return HttpResponse(des_sdate)
            eobj=Employee.objects.create(e_id=id,e_fname=fname,e_lname=lname,e_sname=sname,e_dob=dob,e_bgroup=bgroup,
                                         e_mstatus=mstatus,e_nationality=nationality,e_disb=disb,e_gender=gender,
                                         e_paddr1=paddr1,e_paddr2=paddr2,e_paddr3=paddr3,e_caddr1=caddr1,
                                         e_caddr2=caddr2,e_caddr3=caddr3,e_email=email,e_mnumber=mnumber,e_doj=doj,
                                         e_qfn=qfn,e_prev_exp=prev_exp,e_prev_cmp=prev_cmp,e_cl='2',e_sl='2',e_el='4',
                                         e_fl='2',e_status=status,e_pwd=pwd)
            eobj.save()
            e=Employee.objects.get(e_id=id)
            #return HttpResponse(e.e_id)
            #desobj=Employee_desig.objects.create(ed_eid=id,ed_did=e,ed_sdate=des_sdate,ed_status='Y')
            desobj = Employee_desig.objects.create(ed_eid=Employee.objects.get(e_id=id),
                                                   ed_did=Designation.objects.get(d_id=des_id), ed_sdate=des_sdate,
                                                   ed_status='Y')
            desobj.save()
            salob = Employee_sal.objects.create(es_eid=Employee.objects.get(e_id=id), es_ctc=sal, es_sdate=des_sdate,
                                                es_status='Y')
            salob.save()


            idob.id_val = id + 1
            idob.save()
            logged_in = request.session['eid']
            user = Employee.objects.get(e_id=logged_in)
            dobj = Designation.objects.all().exclude(d_role=0)
            return render(request,'ultimatix/admin/addemp.html',{'success':' Data saved successfully !','dobj':dobj,'user':user})
    else:
        try:
            logged_in = request.session['eid']
        except:
            logged_in = False
        if logged_in:
            user=Employee.objects.get(e_id=logged_in)
            dobj=Designation.objects.all().exclude(d_role=0)
            return render(request,'ultimatix/admin/addemp.html',{'dobj':dobj,'user':user})
        else:
            request.session['eid'] = None
            return redirect('ultimatix:show_login')

def viewemp_view(request):

    #eobj=Employee.objects.all().exclude(e_id='1022890')
    cursor = connection.cursor()
    sql = '''
    select distinct ue.e_id, (ue.e_fname ||' ' ||ue.e_lname) as Name, ue.e_gender, ue.e_doj, ud.d_desc, ues.es_ctc, 
    ue.e_status,ue.e_img,ue.e_email,ue.e_qfn,upm.pid_id
    from ultimatix_employee ue
    left outer join ultimatix_employee_desig ued on ued.ed_eid_id = ue.e_id
    left outer join ultimatix_designation ud on ud.d_id = ued.ed_did_id
    left outer join ultimatix_employee_sal ues on ues.es_eid_id = ue.e_id
    left outer join ultimatix_project_members upm on upm.eid_id = ues.es_eid_id
    where upm.pmedate is null and ued.ed_status = 'Y' and ud.d_cd!='SYSADMIN' and ues.es_status = 'Y' order by (ue.e_fname ||' ' ||ue.e_lname) asc
    '''
    cursor.execute(sql)
    res = cursor.fetchall()
    res_list = []
    for r in res:
        t = {}
        t['eid'] = r[0]
        t['name'] = r[1]
        t['gender']=r[2]
        t['doj']=r[3]
        t['desig']=r[4]
        t['ctc']=r[5]
        t['status']=r[6]
        t['img']=r[7]
        t['email']=r[8]
        t['qfn']=r[9]
        t['pid']=r[10]
        res_list.append(t)
    cursor.close()
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    return render(request,'ultimatix/admin/viewemp.html',{'eobj':res_list,'user':user})


def updtemp_view(request,pk=0):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    if pk!=0:
        eobj=Employee.objects.get(e_id=pk)
        #d1=str(eobj.e_dob)
        dobj = Designation.objects.all().exclude(d_role=0)
        empdesobj=Employee_desig.objects.get(ed_eid=pk,ed_status='Y').ed_did
        empsalobj=Employee_sal.objects.get(es_eid=pk,es_status='Y')
        return render(request, 'ultimatix/admin/updt_emp.html', {'user':user,
                                                                 'eobj': eobj,
                                                                 'dobj':dobj,
                                                                 'empdesobj':empdesobj,
                                                                 'empsalobj':empsalobj})
        #return HttpResponse(pk)
    elif pk==0:
        eid=request.POST.get('eid')
        lname = request.POST.get('lname')
        sname = request.POST.get('surname')
        bgroup = request.POST.get('bgroup')
        mstatus = request.POST.get('mstatus')
        disb = request.POST.get('chk_disb')
        if not disb:
            disb = 'N'
        paddr1 = request.POST.get('paddr1')
        paddr2 = request.POST.get('paddr2')
        paddr3 = request.POST.get('paddr3')
        caddr1 = request.POST.get('caddr1')
        caddr2 = request.POST.get('caddr2')
        caddr3 = request.POST.get('caddr3')
        email = request.POST.get('email')
        mnumber = request.POST.get('mnumber')
        qfn = request.POST.get('qfn')
        des = request.POST.get('desig_select')
        des_id = Designation.objects.get(d_cd=des).d_id
        print(des_id)
        ob = Employee_desig.objects.filter(ed_eid=eid, ed_did=des_id, ed_status='Y')
        if not ob:
            des_sdate = str(datetime.today().strftime('%Y-%m-%d'))
            des_edate = str(datetime.today().strftime('%Y-%m-%d'))
            Employee_desig.objects.filter(ed_eid=eid,ed_status='Y').update(ed_edate=des_edate,ed_status='N')
            desobj = Employee_desig.objects.create(ed_eid=Employee.objects.get(e_id=eid),
                                                   ed_did=Designation.objects.get(d_id=des_id), ed_sdate=des_sdate,
                                                   ed_status='Y')
            desobj.save()
        sal = request.POST.get('salary')
        if(float(sal)!=(Employee_sal.objects.get(es_eid=eid,es_status='Y').es_ctc)):
            sal_sdate = str(datetime.today().strftime('%Y-%m-%d'))
            sal_edate = str(datetime.today().strftime('%Y-%m-%d'))
            Employee_sal.objects.filter(es_eid=eid,es_status='Y').update(es_edate=sal_edate,es_status='N')
            salob = Employee_sal.objects.create(es_eid=Employee.objects.get(e_id=eid), es_ctc=sal, es_sdate=sal_sdate,
                                                es_status='Y')
            salob.save()
            #return HttpResponse('salary updated successfully')
            #return HttpResponse(str(eid)+lname+sname)


        Employee.objects.filter(e_id=eid).update(e_lname=lname,e_sname=sname,e_bgroup=bgroup,
                                                 e_mstatus=mstatus,e_disb=disb,e_paddr1=paddr1,
                                                 e_paddr2=paddr2,e_paddr3=paddr3,e_caddr1=caddr1,
                                                 e_caddr2=caddr2,e_caddr3=caddr3,e_email=email,
                                                 e_mnumber=mnumber,e_qfn=qfn)


        return redirect('ultimatix:view_emp')


def crtprj_view(request):
    if request.method=='POST':
        if request.POST.get('create'):
            title=request.POST.get('tle')
            desc=request.POST.get('desc')
            sdate=request.POST.get('sdate')
            expedate=request.POST.get('expedate')
            #crtd=request.POST.get('crtd')
            #eid=config.eid
            #efname = config.efname
            client=request.POST.get('client')
            cloc=request.POST.get('loct')
            complexity=request.POST.get('size_select')
            type=request.POST.get('type_select')
            budget=request.POST.get('budget')
            if budget:
                budget=float(request.POST.get('budget'))
            else:
                budget=0.0
            logged_in = request.session['eid']
            crtd=Employee.objects.get(e_id=logged_in).e_fname
            if expedate:
                obj=Project.objects.create(ptitle=title,pdesc=desc,psdate=sdate,pexpedate=expedate,pcrtd=crtd,pclient=client,powner=Employee.objects.get(e_id=logged_in),pstatus='A',pcloc=cloc,pcomplexity=complexity,ptype=type,pbudget=budget)
                obj.save()
                id_max = Project.objects.all().aggregate(Max('pid'))['pid__max']
                ob1=Project_members.objects.create(pid=Project.objects.get(pid=id_max),
                                            eid=Employee.objects.get(e_id=logged_in),
                                            rid=Project_roles.objects.get(r_cd='OWN'),
                                            pmsdate=sdate,pmstatus='Y')
                ob1.save()
            else:
                obj = Project.objects.create(ptitle=title, pdesc=desc, psdate=sdate, pcrtd=crtd,pclient=client, powner=Employee.objects.get(e_id=logged_in), pstatus='A',pcloc=cloc,pcomplexity=complexity,ptype=type,pbudget=budget)
                obj.save()
                id_max = Project.objects.all().aggregate(Max('pid'))['pid__max']
                ob1=Project_members.objects.create(pid=Project.objects.get(pid=id_max),
                                            eid=Employee.objects.get(e_id=logged_in),
                                            rid=Project_roles.objects.get(r_cd='OWN'),
                                            pmsdate=sdate,pmstatus='Y')
                ob1.save()

            user = Employee.objects.get(e_id=logged_in)
            return render(request, 'ultimatix/admin/create_project.html', {'user':user,'success':'Data saved successfully'})

            #return HttpResponse('submitted : '+cloc+complexity+type+str(budget))

    else:
        try:
            logged_in = request.session['eid']
        except:
            logged_in = False
        if logged_in:
            user=Employee.objects.get(e_id=logged_in)

            return render(request, 'ultimatix/admin/create_project.html', {'user':user})
        else:
            request.session['eid'] = None
            return redirect('ultimatix:show_login')

def show_login(request):
    if request.method=='POST':
        if request.POST.get('login'):
            eid=request.POST.get('eid')
            pwd=request.POST.get('pwd')
            obj=Employee.objects.filter(e_id=eid,e_pwd=pwd).first()
            if not obj:
                config.eid = ""
                config.efname =""
                request.session['eid'] = None
                return render(request, 'ultimatix/login.html',{'error':'Incorrect Employee Id or Password'})
            else:
                request.session['eid'] = eid
                config.eid = Employee.objects.filter(e_id=eid, e_pwd=pwd).first().e_id
                config.efname=Employee.objects.filter(e_id=eid, e_pwd=pwd).first().e_fname
                return redirect('ultimatix:home_view')

    else:
        config.eid=""
        config.efname =""
        request.session['eid'] = None
        return render(request,'ultimatix/login.html')


def proejcts_view(request):
    try:
        logged_in = request.session['eid']
    except:
        logged_in = False
    if logged_in:
        user = Employee.objects.get(e_id=logged_in)
        pobj=Project.objects.all()
        return render(request,'ultimatix/admin/viewprojects.html',{'pobj':pobj,'user':user})
    else:
        request.session['eid'] = None
        return redirect('ultimatix:show_login')

def updtproj_view(request,pid):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    if pid!=0:
        if request.method=='POST':
            if request.POST.get('update'):
                desc=request.POST.get('desc')
                expedate=request.POST.get('expedate')
                client=request.POST.get('client')
                status = request.POST.get('status')
                if expedate:
                    if status=='C':
                        edate = str(datetime.today().strftime('%Y-%m-%d'))
                        Project.objects.filter(pid=pid).update(pdesc=desc, pexpedate=expedate, pedate=edate,pclient=client,
                                                               pstatus=status)
                        return redirect('ultimatix:view_project')
                    elif status=='A':
                        Project.objects.filter(pid=pid).update(pdesc=desc,pexpedate=expedate,pclient=client,pstatus=status)
                        return redirect('ultimatix:view_project')

                else:
                    if status=='C':
                        edate = str(datetime.today().strftime('%Y-%m-%d'))
                        Project.objects.filter(pid=pid).update(pdesc=desc, pedate=edate,pclient=client,
                                                               pstatus=status)
                        return redirect('ultimatix:view_project')
                    elif status=='A':
                        Project.objects.filter(pid=pid).update(pdesc=desc,pclient=client,pstatus=status)
                        return redirect('ultimatix:view_project')
                    #Project.objects.filter(pid=pid).update(pdesc=desc, pclient=client,pstatus=status)
                    #return redirect('ultimatix:view_project')
        else:

            pobj=Project.objects.get(pid=pid)
            return render(request,'ultimatix/admin/updt_project.html',{'pobj':pobj,'user':user})
    elif pid==0:
        return HttpResponse(pid)


def allocate_view(request):
    if request.is_ajax() and request.GET:
        desid = request.GET.get('desid')
        cursor = connection.cursor()
        if int(desid)!=0:
            #print('desid not 0')
            sql = """
            select e_id,(e_fname||" "||ifnull(e_lname,"")) from ultimatix_employee where e_id in(select es.ed_eid_id from 
            ultimatix_employee_desig es where es.ed_did_id=%s and es.ed_status='Y') and 
            e_status='Y' order by e_id
            """
            cursor.execute(sql,[desid])
        else:
            sql = """
            select e_id,(e_fname||" "||ifnull(e_lname,"")) from ultimatix_employee where e_id in(select es.ed_eid_id from 
            ultimatix_employee_desig es where es.ed_did_id!='16' and es.ed_status='Y') and 
            e_status='Y' order by e_id
            """
            cursor.execute(sql)
        res=cursor.fetchall()
        res_list=[]
        for r in res:
            t={}
            t['eid']=r[0]
            t['ename']=r[1]
            res_list.append(t)
        j=json.dumps(res_list)
        cursor.close()
        #print('called')
        #print('result : '+str(res_list))
        return JsonResponse(res_list, safe=False)
    elif request.method=='POST':
        pid=request.POST.get('pid')
        eid = request.POST.get('eid')
        #print('Employee id: '+eid)
        rid=request.POST.get('role_select')
        desig_cd=Designation.objects.get(d_id=int(request.POST.get('desig_select'))).d_cd
        sdate=str(datetime.now())

        ###if project is not yet started
        #pdate=Project.objects.get()

        #### if employee is assigned to same project already
        cursor = connection.cursor()
        sql = """
        select eid_id from ultimatix_project_members where eid_id=%s 
        and pmstatus='Y'
        """
        cursor.execute(sql,[eid])
        res=cursor.fetchall()
        #print('result : '+str(res))

        #### if project have an active hr
        sql='''
        select ue.e_id
        from ultimatix_employee ue
        left outer join ultimatix_project_members upm on upm.eid_id = ue.e_id
        left outer join ultimatix_project_roles up on up.r_id = upm.rid_id
        where upm.pmstatus = 'Y' and up.r_cd='MNG' and ue.e_status = 'Y' and upm.pid_id=%s
        '''
        cursor.execute(sql, [pid])
        res1=cursor.fetchall()

        sql='''
        select ue.e_id
        from ultimatix_employee ue
        left outer join ultimatix_project_members upm on upm.eid_id = ue.e_id
        left outer join ultimatix_project_roles up on up.r_id = upm.rid_id
        where upm.pmstatus = 'Y' and up.r_cd='PM' and ue.e_status = 'Y' and upm.pid_id=%s        
        '''
        cursor.execute(sql, [pid])
        res2=cursor.fetchall()


        cursor.close()
        if res:
            #if employee is assigned to same project already
            print('case 1')
            logged_in = request.session['eid']
            user = Employee.objects.get(e_id=logged_in)
            result=default_allocate_view()
            return render(request,'ultimatix/admin/allocate.html',{'pobj': result[0], 'dobj': result[1], 'eobj': result[2], 'robj': result[3],'error':' Employee is already allocated !','user':user})
        elif not res1 and (desig_cd not in ('HR','HR-M')):
            #if project does not have an HR nad currently not assigning an HR
            print('case 2')
            logged_in = request.session['eid']
            user = Employee.objects.get(e_id=logged_in)
            #print('role id : '+str(rid))
            result = default_allocate_view()
            return render(request, 'ultimatix/admin/allocate.html',
                          {'pobj': result[0], 'dobj': result[1], 'eobj': result[2], 'robj': result[3],
                           'error': ' Assign an HR first !','user':user})
            '''elif not res2 and desig_cd not in ('PM'):
            print('case 3')
            print(str(res1)+desig_cd)
            logged_in = request.session['eid']
            user = Employee.objects.get(e_id=logged_in)
            result = default_allocate_view()
            return render(request, 'ultimatix/admin/allocate.html',
                          {'pobj': result[0], 'dobj': result[1], 'eobj': result[2], 'robj': result[3],
                           'error': ' Please assign a manager for the project !','user':user})'''
            '''elif res1:
            # if project already have an HR
            logged_in = request.session['eid']
            user = Employee.objects.get(e_id=logged_in)
            result = default_allocate_view()
            return render(request, 'ultimatix/admin/allocate.html',
                          {'pobj': result[0], 'dobj': result[1], 'eobj': result[2], 'robj': result[3],
                           'error': ' Project already has an HR !','user':user})'''
        else:

            ob = Project_members.objects.create(pid=Project.objects.get(pid=pid),
                                            eid=Employee.objects.get(e_id=eid),
                                            rid=Project_roles.objects.get(r_id=rid),
                                            pmsdate=sdate,pmstatus='Y')
            ob.save()
            logged_in = request.session['eid']
            user = Employee.objects.get(e_id=logged_in)
            result=default_allocate_view()
            return render(request,'ultimatix/admin/allocate.html',{'pobj': result[0], 'dobj': result[1], 'eobj': result[2], 'robj': result[3],'success':' Data saved successfully !','user':user})
    else:
        try:
            logged_in = request.session['eid']
        except:
            logged_in = False
        if logged_in:
            user=Employee.objects.get(e_id=logged_in)
            result=default_allocate_view()
            return render(request,'ultimatix/admin/allocate.html',{'pobj': result[0], 'dobj': result[1], 'eobj': result[2], 'robj': result[3],'user':user})
        else:
            request.session['eid'] = None
            return redirect('ultimatix:show_login')

def default_allocate_view():
    s = datetime.now().date()
    #pobj = Project.objects.all().exclude(pstatus='C')
    pobj=Project.objects.all().exclude(psdate__gt=s).exclude(pstatus='C')
    dobj = Designation.objects.all().exclude(d_cd='SYSADMIN')
    initial_desid=Designation.objects.all().exclude(d_cd='SYSADMIN').first().d_id
    cursor = connection.cursor()
    sql = """
    select e_id,(e_fname||" "||ifnull(e_lname,"")) from ultimatix_employee where e_id in(select es.ed_eid_id from 
    ultimatix_employee_desig es where es.ed_did_id=%s and es.ed_status='Y') and 
    e_status='Y' order by e_id
    """
    cursor.execute(sql,[initial_desid])
    res = cursor.fetchall()
    res_list = []
    for r in res:
        t={}
        #s=str(r[0])+","+str(r[1])
        #t['eid']=r[0]
        #t['ename']=r[1]
        #t.append(s)
        t['eid'] = (r[0])
        t['ename'] = (r[1])
        #t=r[0]
        res_list.append(t)
    eobj = json.dumps(res_list)
    #print('eobj:'+str(eobj))
    robj = Project_roles.objects.all().exclude(r_active='N').exclude(r_cd='OWN')
    result=[pobj,dobj,res_list,robj]
    cursor.close()
    return result


def proejctmembers_view(request,pid):
    if pid:
        if request.method=='POST':
            #pmid=request.POST.get('pmid')
            id=request.POST.get('submit')
            pmid=request.POST.get(request.POST.get('submit')+"_pmid")
            pmedate = str(datetime.now())
            #edate=(datetime.now().strftime('%Y-%m-%d  %H:%M:%S'))
            Project_members.objects.filter(pmid=pmid).update(pmedate=pmedate, pmstatus='N')
            p = Project.objects.get(pid=pid)
            test = Project_members.objects.filter(pid=pid)
            cursor = connection.cursor()
            sql = """
            SELECT pm.pmid,e.e_id, (e.e_fname||" "||ifnull(e.e_lname,"")), pr.r_desc, pm.pmsdate, pm.pmedate, pm.pmstatus
            FROM ultimatix_project p
            LEFT OUTER JOIN ultimatix_project_members pm ON pm.pid_id = p.pid
            LEFT OUTER JOIN ultimatix_project_roles pr ON pr.r_id = pm.rid_id
            LEFT OUTER JOIN ultimatix_employee e ON e.e_id = pm.eid_id
            WHERE p.pid=%s order by pr.r_id DESC 
            """
            cursor.execute(sql, [pid])
            res = cursor.fetchall()
            res_list = []
            for r in res:
                t = {}
                t['pmid'] = str(r[0])
                t['eid'] = str(r[1])
                t['name'] = str(r[2])
                t['prdesc'] = str(r[3])
                #t['pmsdate'] = str((r[4]).strftime('%Y-%m-%d  %H:%M:%S'))
                str1 = str(r[4])
                str1 = str1.split(".")
                t['pmsdate'] = str1[0]
                '''if not str(r[5]):
                    t['pmedate'] = str((r[5]).strftime('%Y-%m-%d  %H:%M:%S'))
                else:
                    t['pmedate'] = str(r[5])'''
                if r[5] is not None:
                    str1=str(r[5])
                    str1=str1.split(".")
                    t['pmedate']=str1[0]
                else:
                    t['pmedate']="-"
                t['pmstatus'] = str(r[6])
                res_list.append(t)
            logged_in = request.session['eid']
            user = Employee.objects.get(e_id=logged_in)
            return render(request, 'ultimatix/admin/updtprojectmembers.html', {'p': p, 'res': res_list, 'user': user})

            #return HttpResponse('success')
        else:
            p=Project.objects.get(pid=pid)
            test = Project_members.objects.filter(pid=pid)
            #print(test)
            #pobj=Project_members.objects.filter(pid=pid)
            cursor = connection.cursor()
            sql = """
            SELECT pm.pmid,e.e_id, (e.e_fname||" "||ifnull(e.e_lname,"")), pr.r_desc, pm.pmsdate, pm.pmedate, pm.pmstatus
            FROM ultimatix_project p
            LEFT OUTER JOIN ultimatix_project_members pm ON pm.pid_id = p.pid
            LEFT OUTER JOIN ultimatix_project_roles pr ON pr.r_id = pm.rid_id
            LEFT OUTER JOIN ultimatix_employee e ON e.e_id = pm.eid_id
            WHERE p.pid=%s order by pr.r_id DESC 
            """
            cursor.execute(sql,[pid])
            res = cursor.fetchall()
            res_list = []
            for r in res:
                t={}
                t['pmid']=str(r[0])
                t['eid']=str(r[1])
                t['name']=str(r[2])
                t['prdesc']=str(r[3])
                #t['pmsdate']=str((r[4]).strftime('%Y-%m-%d  %H:%M:%S'))
                str1 = str(r[4])
                str1 = str1.split(".")
                t['pmsdate'] = str1[0]
                if r[5] is not None:
                    str1=str(r[5])
                    str1=str1.split(".")
                    t['pmedate']=str1[0]
                else:
                    t['pmedate']="-"

                t['pmstatus']=str(r[6])
                res_list.append(t)
            logged_in = request.session['eid']
            user = Employee.objects.get(e_id=logged_in)
            #print(str(res_list))
            return render(request,'ultimatix/admin/updtprojectmembers.html',{'p':p,'res':res_list,'user':user})

    else:
        raise Http404


def mark_attendance(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    if request.method=='POST':
        if request.POST.get('validate'):
            adate=request.POST.get('adate')
            fname=request.FILES["myFile"]
            ext = os.path.splitext(fname.name)[1]
            if ext.lower() in ['.xlsx']:
                wb = openpyxl.load_workbook(fname)
                worksheet = wb.worksheets[0]
                try:
                    excel_list=[]
                    for i in range(1,worksheet.max_row+1):
                        if i==1:
                            continue
                        data_list={}
                        data_list['eid']=str(worksheet.cell(row=i,column=1).value)
                        data_list['swipes']=str(worksheet.cell(row=i,column=2).value)
                        ip = str(worksheet.cell(row=i, column=2).value)
                        temp = ip.split(',')
                        l = len(temp)
                        if l % 2 != 0:
                            temp.pop(-1)
                        temp_time_hldr = []
                        for i in temp:
                            datetime_object = datetime.strptime(i, '%H:%M')
                            temp_time_hldr.append(datetime_object)
                            odd_place = temp_time_hldr[::2]
                            even_place = temp_time_hldr[1:][::2]
                            total = []
                            for i, j in zip(even_place, odd_place):
                                diff = i - j
                                total.append(diff)
                            total_time = datetime(1900, 1, 1, 0, 0)
                            for i in total:
                                total_time = total_time + i
                        data_list['total_working']=total_time.strftime('%H:%M')
                        hours, minutes = (total_time.strftime('%H:%M')).split(':')
                        tot=int(hours) * 60 + int(minutes)
                        if tot>=(9*60):
                            data_list['attendance']='P'
                        elif tot>=(4.5*60):
                            data_list['attendance'] = 'H'
                        else:
                            data_list['attendance'] = 'A'
                        excel_list.append(data_list)
                        #print('adate : '+adate)
                        temp=Timesheet.objects.filter(eid=Employee.objects.get(e_id=data_list['eid']),tdate=adate)
                        if not temp:
                            ob=Timesheet.objects.create(eid=Employee.objects.get(e_id=data_list['eid']),tdate=adate,thours=data_list['total_working'],tstate=data_list['attendance'])
                            ob.save()
                        else:
                            Timesheet.objects.filter(eid=Employee.objects.get(e_id=data_list['eid']), tdate=adate).update(thours=data_list['total_working'],tstate=data_list['attendance'])

                    return render(request,'ultimatix/upload_attendance.html',{'adate':adate,'excel_list':excel_list,'user':user})
                except Exception as e:
                    print('Exception : '+str(e))
                    return render(request, 'ultimatix/attendance.html', {'error': ' Some error occured while processing the data !','e':e,'user':user})
                #return redirect('ultimatix:upload_attendance', adate)
            else:
                return render(request, 'ultimatix/attendance.html',{'error':' Invalid File Format !','user':user})
        elif request.POST.get('submit'):
            return HttpResponse('sumbit')
    else:
        return render(request,'ultimatix/attendance.html',{'user':user})



def upload_attendance(request,adate):
    return render(request,'ultimatix/upload_attendance.html',{'adate':adate})


def apply_leave(request):
    if request.method=='POST':
        if request.POST.get('submit'):
            eid=request.session['eid']
            sdate=datetime.strptime(request.POST.get('sdate'), '%m/%d/%Y').strftime('%Y-%m-%d')
            edate = datetime.strptime(request.POST.get('edate'), '%m/%d/%Y').strftime('%Y-%m-%d')
            nof=float(request.POST.get('nof'))
            ltype=request.POST.get('leave_select')
            desc = request.POST.get('rsn')
            status='P'
            reqdate = str(datetime.now().strftime('%Y-%m-%d'))
            queueid = findqueueid(eid)

            Leave.objects.create(eid=Employee.objects.get(e_id=eid),sdate=sdate,edate=edate,nof=nof,
                                 ltype=ltype,desc=desc,status=status,reqdate=reqdate,queueid=queueid)
            if ltype=='0':
                rem=(Employee.objects.get(e_id=eid).e_cl)-nof
                Employee.objects.filter(e_id=eid).update(e_cl=rem)
            elif ltype=='1':
                rem=(Employee.objects.get(e_id=eid).e_sl)-nof
                Employee.objects.filter(e_id=eid).update(e_sl=rem)
            elif ltype=='2':
                rem=(Employee.objects.get(e_id=eid).e_el)-nof
                Employee.objects.filter(e_id=eid).update(e_el=rem)
            elif ltype=='3':
                rem=(Employee.objects.get(e_id=eid).e_fl)-nof
                Employee.objects.filter(e_id=eid).update(e_fl=rem)

            logged_in = request.session['eid']
            lobj = Leave.objects.filter(eid=logged_in)
            print('going queue : '+str(queueid))
            logged_in = request.session['eid']
            user=Employee.objects.get(e_id=logged_in)
            return render(request,'ultimatix/applyleave.html',{'user':user,'success':'Leave request submitted !','lobj':lobj})


            #return HttpResponse(queueid)
    else:
        try:
            logged_in = request.session['eid']
        except:
            logged_in = False
        if logged_in:
            user=Employee.objects.get(e_id=logged_in)
            lobj=Leave.objects.filter(eid=logged_in)
            return render(request,'ultimatix/applyleave.html',{'user':user,'lobj':lobj})
        else:
            request.session['eid'] = None
            return redirect('ultimatix:show_login')

def findqueueid(eid):
    ob=Project_members.objects.filter(eid_id=eid,pmstatus='Y')
    if not ob:
        print('unallocated going to admin : 1022890')
        return '1022890'
    title=Project.objects.get(pid=Project_members.objects.get(eid_id=eid,pmstatus='Y').pid_id).ptitle
    #print('title: '+title)
    if title=='TCS - MNG':
        print('going to admin : 1022890')
        return '1022890'
    else:
        pid= Project_members.objects.get(eid_id=eid,pmstatus='Y').pid_id
        hreid=Project_members.objects.get(rid_id=Project_roles.objects.get(r_cd='MNG').r_id,pid_id=pid, pmstatus='Y').eid_id
        #print(str(hreid)+" "+str(eid))
        if str(hreid)==str(eid):
            hrmeid=Employee_desig.objects.get(ed_did_id=Designation.objects.get(d_cd='HR-M').d_id,
                                              ed_status='Y').ed_eid_id
            print('going to hr manager : '+str(hrmeid))
            return hrmeid
        else:
            pmeid=Project_members.objects.get(rid_id=Project_roles.objects.get(r_cd='PM').r_id, pid_id=pid,
                                        pmstatus='Y').eid_id
            if str(pmeid)==str(eid):
                print('going to hr : '+str(hreid))
                return hreid
            else:
                print('going to pm : '+str(pmeid))
                return pmeid


def leave_request(request):
    if request.method=='POST':
        logged_in = request.session['eid']
        if request.POST.get('submit'):
            lid=request.POST.get('submit')
            rem=request.POST.get(request.POST.get('submit')+"_remarks")
            print('remarks :'+rem)
            Leave.objects.filter(lid=lid).update(status='A',updtby=logged_in,remarks=rem)
            user = Employee.objects.get(e_id=logged_in)
            #lreqobj=Leave.objects.filter(queueid=logged_in)
            lreqobj = Leave.objects.filter(queueid=logged_in).order_by('-reqdate')
            return render(request, 'ultimatix/leaverequest.html', {'user': user,'lreqobj':lreqobj,'success':' Data saved successfully !'})

        elif request.POST.get('cancel'):
            lid = request.POST.get('cancel')
            rem = request.POST.get(request.POST.get('cancel') + "_remarks")

            nof=float(Leave.objects.get(lid=lid).nof)
            ltype=Leave.objects.get(lid=lid).ltype
            eid=Leave.objects.get(lid=lid).eid_id
            r=-1.0
            if ltype == 0:
                r = (Employee.objects.get(e_id=eid).e_cl) + nof
                Employee.objects.filter(e_id=eid).update(e_cl=r)
            elif ltype == 1:
                r = (Employee.objects.get(e_id=eid).e_sl) + nof
                Employee.objects.filter(e_id=eid).update(e_sl=r)
            elif ltype == 2:
                r = (Employee.objects.get(e_id=eid).e_el) + nof
                Employee.objects.filter(e_id=eid).update(e_el=r)
            elif ltype == 3:
                r = (Employee.objects.get(e_id=eid).e_fl) + nof
                Employee.objects.filter(e_id=eid).update(e_fl=r)


            Leave.objects.filter(lid=lid).update(status='R', updtby=logged_in,remarks=rem)
            print('remarks :' + rem)
            user = Employee.objects.get(e_id=logged_in)
            lreqobj = Leave.objects.filter(queueid=logged_in).order_by('-reqdate')
            return render(request, 'ultimatix/leaverequest.html', {'user': user,'lreqobj':lreqobj,'success':' Data saved successfully !'})

        elif request.POST.get('fwd'):
            logged_in = request.session['eid']
            user = Employee.objects.get(e_id=logged_in)
            lreqobj = Leave.objects.filter(queueid=logged_in).order_by('-reqdate')
            lid=request.POST.get('fwd')
            fwdto=str(request.POST.get(request.POST.get('fwd')+"_fwdto"))
            print('fwdto : '+fwdto)
            #fwdto=str(request.POST.get(''))

            if fwdto=="":
                return render(request, 'ultimatix/leaverequest.html', {'user': user, 'lreqobj': lreqobj,'error':'Please enter employee id'})
                #return HttpResponse('Please enter en employee id'+lid+" "+str(fwdto))
            else:
                ob=Employee.objects.filter(e_id=fwdto,e_status='Y')
                if not ob:
                    return render(request, 'ultimatix/leaverequest.html',
                                  {'user': user, 'lreqobj': lreqobj, 'error': 'Please enter a valid employee id'})
                    #return HttpResponse('please enter a valid employee id'+lid+" "+str(fwdto))
                else:
                    ownid=Leave.objects.get(lid=lid).eid_id
                    if(str(fwdto)==str(ownid)):
                        return render(request, 'ultimatix/leaverequest.html',
                                      {'user': user, 'lreqobj': lreqobj, 'error': ' Invalid Employee !'})
                        #return HttpResponse('Invalid Employee'+lid+" "+str(fwdto))
                    else:
                        lid = request.POST.get('fwd')
                        rem=request.POST.get(request.POST.get('fwd') + "_remarks")
                        queueid=fwdto
                        #updt=str(request.session['eid'])
                        Leave.objects.filter(lid=lid).update(queueid=queueid,updtby=logged_in,
                                                             remarks=rem)
                        return render(request, 'ultimatix/leaverequest.html',
                                      {'user': user, 'lreqobj': lreqobj, 'success': ' Data saved successfully !'})
                        #return HttpResponse('success : '+lid+" "+rem+" "+queueid+" "+updt)
    else:
        try:
            logged_in = request.session['eid']
        except:
            logged_in = False
        if logged_in:
            user = Employee.objects.get(e_id=logged_in)
            lreqobj=Leave.objects.filter().order_by('-reqdate')
            return render(request, 'ultimatix/leaverequest.html', {'user': user,'lreqobj':lreqobj})
        else:
            request.session['eid'] = None
            return redirect('ultimatix:show_login')

def generate_payslip(request):
    if request.method=='POST':
        logged_in = request.session['eid']
        user = Employee.objects.get(e_id=logged_in)
        adate=request.POST.get('sdate')
        r=Wage.objects.filter(eid_id=logged_in,sdate__lte=adate,
                              edate__gte=adate,status='Y')
        cursor = connection.cursor()
        if r:
            sql = '''
            select ue.e_id, (ue.e_fname ||' ' ||ifnull(ue.e_lname,'')) as Name, ud.d_desc, ues.es_ctc
            from ultimatix_employee ue
            left outer join ultimatix_employee_desig ued on ued.ed_eid_id = ue.e_id
            left outer join ultimatix_designation ud on ud.d_id = ued.ed_did_id
            left outer join ultimatix_employee_sal ues on ues.es_eid_id = ue.e_id
            where ued.ed_status = 'Y' and ues.es_status = 'Y' and ue.e_id = %s
            '''
            cursor.execute(sql,[logged_in])
            res=cursor.fetchone()
            t={}
            t['eid']=res[0]
            t['name']=res[1]
            t['desig']=res[2]
            t['ctc']=res[3]


            sql='''
            select bs,conv,hra,city,sundry,ptax,pf,esis,lop from ultimatix_wage where eid_id=%s and 
            status='Y' and sdate<=%s and edate>=%s
            '''
            cursor.execute(sql, [logged_in,adate,adate])
            res = cursor.fetchone()
            t['bs']=res[0]
            t['conv']=res[1]
            t['hra']=res[2]
            t['city']=res[3]
            t['sundry']=res[4]
            t['ptax']=res[5]
            t['pf']=res[6]
            t['esis']=res[7]
            t['lop']=res[8]
            t['earn']=float(res[0])+float(res[1])+float(res[2])+float(res[3])+float(res[4])
            t['ded']=float(res[5])+float(res[6])+float(res[7])
            t['fin']=round((float(res[0])+float(res[1])+float(res[2])+float(res[3])+float(res[4]))-float(res[8]),2)
            cursor.close()

            date=datetime.strptime(adate,'%Y-%m-%d')
            mnyr = date.strftime('%B') + ", " + date.strftime('%Y')
            t['mnyr']=mnyr
            pdf = render_to_pdf('pdf/payslip.html', t)
            return HttpResponse(pdf, content_type='application/pdf')
            #return HttpResponse('record present')
        else:
            return render(request, 'ultimatix/payslip.html', {'user': user,'error':'Payslip is not yet generated !'})
            #return HttpResponse('record not present')
        '''cursor = connection.cursor()
        sql='''

        '''
        cursor.execute(sql)
        res=cursor.fetchone()
        #print('res[0]:'+ str(res[0]))
        for r in res:
            t={}
            t['lid']=str(res[0])
            t['sdate']=str(res[1])
            t['edate']=str(res[2])
        pdf = render_to_pdf('pdf/payslip.html', t)
        cursor.close()
        return HttpResponse(pdf, content_type='application/pdf')'''

    else:
        try:
            logged_in = request.session['eid']
        except:
            logged_in = False
        if logged_in:
            user = Employee.objects.get(e_id=logged_in)
            print('inside salary view')
            return render(request, 'ultimatix/payslip.html', {'user': user})
        else:
            request.session['eid'] = None
            return redirect('ultimatix:show_login')


def calendar_view(request):
    logged_in = request.session['eid']
    user=Employee.objects.get(e_id=logged_in)
    if request.method=='POST':
        if request.POST.get('submit'):
            month=(request.POST.get('month_select'))
            m=month
            year =(request.POST.get('year_select'))
            eid=request.POST.get('emp_select')
            ed = str(calendar.monthrange(int(year), int(month))[1])
            if len(month) == 1:
                month = '0' + month
            sdate = year + "-" + month + "-01"
            edate = year + "-" + month + "-" + ed

            cal = calendar.Calendar()
            hobj = Holidays.objects.filter(hdate__gte=sdate, hdate__lte=edate)
            holidays = []
            for o in hobj:
                holidays.append(o.hdate.day)
            tobj = Timesheet.objects.filter(eid=eid, tdate__gte=sdate, tdate__lte=edate)
            absent = []
            halfday = []
            for o in tobj:
                if o.tstate == 'A':
                    absent.append(o.tdate.day)
                elif o.tstate == 'H':
                    halfday.append(o.tdate.day)
            calendar_list = []
            for day in cal.itermonthdays(int(year), int(month)):
                t = {}
                t['date'] = day
                if day in holidays:
                    t['value'] = 0
                elif day in absent:
                    t['value'] = 1
                elif day in halfday:
                    t['value'] = 2
                else:
                    t['value'] = 3
                calendar_list.append(t)
            #emplist = Employee.objects.filter(~Q(e_id='1022890'))
            emplist = Employee.objects.all()
            return render(request, 'ultimatix/calendar.html',
                          {'user': user, 'calendar_list': calendar_list, 'emplist': emplist,'hobj': hobj,'month':m,'year':year,'eid':eid})
            #return HttpResponse(m+year+eid)
    else:
        calendar_list=[]
        #emplist=Employee.objects.filter(~Q(e_id='1022890'))
        emplist=Employee.objects.all()
        return render(request,'ultimatix/calendar.html',{'user':user,'calendar_lidt':calendar_list,'emplist':emplist})

'''def deallocate_view(request,pmid):
    if pmid:
        return render(request,'ultimatix/admin/deallocate.html',{'pmid':pmid})'''


def news_feed(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    if request.method=='POST':
        head=request.POST.get('head')
        desc=request.POST.get('desc')
        img=request.FILES['myfile']
        News.objects.create(head=head,desc=desc,img=img)
        return render(request, 'ultimatix/admin/news.html', {'success': 'News added successfully !','user':user})
    else:
        return render(request,'ultimatix/admin/news.html',{'user':user})


def empview_projects(request):
    try:
        logged_in = request.session['eid']
    except:
        logged_in = False
    if logged_in:
        r=findrole(logged_in)
        user = Employee.objects.get(e_id=logged_in)
        pobj=Project.objects.all()
        return render(request,'ultimatix/employee/viewprojects.html',{'pobj':pobj,'user':user,'role':r})
    else:
        request.session['eid'] = None
        return redirect('ultimatix:show_login')

def findrole(logged_in):
    cursor=connection.cursor()
    sql = '''
    select ud.d_role
    from ultimatix_employee ue
    left outer join ultimatix_employee_desig ued on ued.ed_eid_id = ue.e_id
    left outer join ultimatix_designation ud on ud.d_id = ued.ed_did_id
    where ued.ed_status = 'Y' and ue.e_id = %s
    '''
    cursor.execute(sql, [logged_in])
    res = cursor.fetchone()
    if str(res[0]) == '0':
        role = 0
    elif str(res[0]) == '2':
        role = 2
    elif str(res[0]) == '1':
        role = 1
    cursor.close()
    return role

def empproejctmembers_view(request,pid):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    if pid:
        cursor = connection.cursor()
        p = Project.objects.get(pid=pid)
        sql = """
        SELECT pm.pmid,e.e_id, (e.e_fname||" "||ifnull(e.e_lname,"")), pr.r_desc, pm.pmsdate, pm.pmedate, pm.pmstatus
        FROM ultimatix_project p
        LEFT OUTER JOIN ultimatix_project_members pm ON pm.pid_id = p.pid
        LEFT OUTER JOIN ultimatix_project_roles pr ON pr.r_id = pm.rid_id
        LEFT OUTER JOIN ultimatix_employee e ON e.e_id = pm.eid_id
        WHERE p.pid=%s order by pr.r_id DESC 
        """
        cursor.execute(sql, [pid])
        res = cursor.fetchall()
        res_list = []
        for r in res:
            t = {}
            t['pmid'] = str(r[0])
            t['eid'] = str(r[1])
            t['name'] = str(r[2])
            t['prdesc'] = str(r[3])
            t['pmsdate'] = str((r[4]).strftime('%Y-%m-%d  %H:%M:%S'))
            if not str(r[5]):
                t['pmedate'] = str((r[5]).strftime('%Y-%m-%d  %H:%M:%S'))
            else:
                t['pmedate'] = str(r[5])
            t['pmstatus'] = str(r[6])
            res_list.append(t)
        r = findrole(logged_in)
        cursor.close()
        return render(request,'ultimatix/employee/viewprojectmembers.html',{'p':p,'res':res_list,'user':user,'role':r})
    else:
        raise Http404


def empmyproject(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    ob=Project_members.objects.filter(eid_id=logged_in, pmstatus='Y')
    if ob:
        print('allocated')
        pid=Project_members.objects.get(eid_id=logged_in, pmstatus='Y').pid_id
        prole=Project_members.objects.get(eid_id=logged_in, pmstatus='Y')
        proledesc=Project_roles.objects.get(r_id=Project_members.objects.get(eid_id=logged_in, pmstatus='Y').rid_id).r_desc
        cursor = connection.cursor()
        p = Project.objects.get(pid=pid)
        sql = """
        SELECT pm.pmid,e.e_id, (e.e_fname||" "||ifnull(e.e_lname,"")), pr.r_desc, pm.pmsdate, pm.pmedate, pm.pmstatus
        FROM ultimatix_project p
        LEFT OUTER JOIN ultimatix_project_members pm ON pm.pid_id = p.pid
        LEFT OUTER JOIN ultimatix_project_roles pr ON pr.r_id = pm.rid_id
        LEFT OUTER JOIN ultimatix_employee e ON e.e_id = pm.eid_id
        WHERE p.pid=%s order by pr.r_id DESC 
        """
        cursor.execute(sql, [pid])
        res = cursor.fetchall()
        res_list = []
        for r in res:
            t = {}
            t['pmid'] = str(r[0])
            t['eid'] = str(r[1])
            t['name'] = str(r[2])
            t['prdesc'] = str(r[3])
            t['pmsdate'] = str((r[4]).strftime('%Y-%m-%d  %H:%M:%S'))
            if not str(r[5]):
                t['pmedate'] = str((r[5]).strftime('%Y-%m-%d  %H:%M:%S'))
            else:
                t['pmedate'] = str(r[5])
            t['pmstatus'] = str(r[6])
            res_list.append(t)

        sql='''
        SELECT p.pid,p.pdesc,e.e_id, (e.e_fname||" "||ifnull(e.e_lname,"")), pr.r_desc, pm.pmsdate, pm.pmedate, pm.pmstatus
        FROM ultimatix_project p
        LEFT OUTER JOIN ultimatix_project_members pm ON pm.pid_id = p.pid
        LEFT OUTER JOIN ultimatix_project_roles pr ON pr.r_id = pm.rid_id
        LEFT OUTER JOIN ultimatix_employee e ON e.e_id = pm.eid_id
        WHERE e.e_id=%s and pm.pmstatus='N' order by pm.pmsdate DESC         
        '''
        cursor.execute(sql, [logged_in])
        res2 = cursor.fetchall()
        res_list2 = []
        for r in res2:
            t = {}
            t['pid'] = str(r[0])
            t['pdesc'] = str(r[1])
            t['ename'] = str(r[3])
            t['rdesc'] = str(r[4])
            t['pmsdate'] = str((r[5]).strftime('%Y-%m-%d  %H:%M:%S'))
            t['pmedate'] = str((r[6]).strftime('%Y-%m-%d  %H:%M:%S'))
            res_list2.append(t)
        #print('role : '+str(role))
        cursor.close()
        return render(request, 'ultimatix/employee/myproject.html',
                      {'p':p,'res':res_list,'res2':res_list2,'user': user, 'role': role,'prole':prole,'proledesc':proledesc})
    else:
        cursor = connection.cursor()
        sql = '''
        SELECT p.pid,p.pdesc,e.e_id, (e.e_fname||" "||ifnull(e.e_lname,"")), pr.r_desc, pm.pmsdate, pm.pmedate, pm.pmstatus
        FROM ultimatix_project p
        LEFT OUTER JOIN ultimatix_project_members pm ON pm.pid_id = p.pid
        LEFT OUTER JOIN ultimatix_project_roles pr ON pr.r_id = pm.rid_id
        LEFT OUTER JOIN ultimatix_employee e ON e.e_id = pm.eid_id
        WHERE e.e_id=%s and pm.pmstatus='N' order by pm.pmsdate DESC         
        '''
        cursor.execute(sql, [logged_in])
        res2 = cursor.fetchall()
        res_list2 = []
        for r in res2:
            t = {}
            t['pid'] = str(r[0])
            t['pdesc'] = str(r[1])
            t['ename'] = str(r[3])
            t['rdesc'] = str(r[4])
            t['pmsdate'] = str((r[5]).strftime('%Y-%m-%d  %H:%M:%S'))
            t['pmedate'] = str((r[6]).strftime('%Y-%m-%d  %H:%M:%S'))
            res_list2.append(t)
        cursor.close()
        return render(request, 'ultimatix/employee/myproject.html',
                      {'user': user, 'role': role,'res2':res_list2,'msg':' You are not allocated !'})

def empdetails(request):
    cursor = connection.cursor()
    sql = '''
    select distinct ue.e_id, (ue.e_fname ||' ' ||ue.e_lname) as Name, ue.e_gender, ue.e_doj, ud.d_desc, ues.es_ctc, 
    ue.e_status,ue.e_img,ue.e_email,ue.e_qfn,upm.pid_id
    from ultimatix_employee ue
    left outer join ultimatix_employee_desig ued on ued.ed_eid_id = ue.e_id
    left outer join ultimatix_designation ud on ud.d_id = ued.ed_did_id
    left outer join ultimatix_employee_sal ues on ues.es_eid_id = ue.e_id
    left outer join ultimatix_project_members upm on upm.eid_id = ues.es_eid_id
    where upm.pmedate is null and ued.ed_status = 'Y' and ud.d_cd!='SYSADMIN' and ues.es_status = 'Y' order by (ue.e_fname ||' ' ||ue.e_lname) asc
    '''
    cursor.execute(sql)
    res = cursor.fetchall()
    res_list = []
    for r in res:
        t = {}
        t['eid'] = r[0]
        t['name'] = r[1]
        t['gender']=r[2]
        t['doj']=r[3]
        t['desig']=r[4]
        t['ctc']=r[5]
        t['status']=r[6]
        t['img']=r[7]
        t['email']=r[8]
        t['qfn']=r[9]
        t['pid']=r[10]
        res_list.append(t)
    cursor.close()
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    print('res_list : '+str(res_list))
    return render(request,'ultimatix/employee/viewemp.html',{'eobj':res_list,'role':role,'user':user})

def empleave(request):
    logged_in = request.session['eid']
    role=findrole(logged_in)
    if request.method=='POST':
        eid = request.session['eid']
        sdate = datetime.strptime(request.POST.get('sdate'), '%m/%d/%Y').strftime('%Y-%m-%d')
        edate = datetime.strptime(request.POST.get('edate'), '%m/%d/%Y').strftime('%Y-%m-%d')
        nof = float(request.POST.get('nof'))
        ltype = request.POST.get('leave_select')
        desc = request.POST.get('rsn')
        status = 'P'
        reqdate = str(datetime.now().strftime('%Y-%m-%d'))
        queueid = findqueueid(eid)

        Leave.objects.create(eid=Employee.objects.get(e_id=eid), sdate=sdate, edate=edate, nof=nof,
                             ltype=ltype, desc=desc, status=status, reqdate=reqdate, queueid=queueid)
        if ltype == '0':
            rem = (Employee.objects.get(e_id=eid).e_cl) - nof
            Employee.objects.filter(e_id=eid).update(e_cl=rem)
        elif ltype == '1':
            rem = (Employee.objects.get(e_id=eid).e_sl) - nof
            Employee.objects.filter(e_id=eid).update(e_sl=rem)
        elif ltype == '2':
            rem = (Employee.objects.get(e_id=eid).e_el) - nof
            Employee.objects.filter(e_id=eid).update(e_el=rem)
        elif ltype == '3':
            rem = (Employee.objects.get(e_id=eid).e_fl) - nof
            Employee.objects.filter(e_id=eid).update(e_fl=rem)

        lobj = Leave.objects.filter(eid=logged_in).order_by('-lid')
        user = Employee.objects.get(e_id=logged_in)
        return render(request, 'ultimatix/employee/applyleave.html', {'user': user, 'lobj': lobj, 'role': role,'success':'Leave request submitted !'})
    else:
        lobj = Leave.objects.filter(eid=logged_in).order_by('-lid')
        user = Employee.objects.get(e_id=logged_in)
        return render(request, 'ultimatix/employee/applyleave.html', {'user': user, 'lobj': lobj,'role':role})


def empleavereq(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    if request.method=='POST':
        if request.POST.get('submit'):
            lid = request.POST.get('submit')
            rem = request.POST.get(request.POST.get('submit') + "_remarks")
            Leave.objects.filter(lid=lid).update(status='A', updtby=logged_in, remarks=rem)
            #print(lid)
            lreqobj = Leave.objects.filter(queueid=logged_in).order_by('-reqdate')
            return render(request, 'ultimatix/employee/leaverequest.html', {'user': user, 'role': role, 'lreqobj': lreqobj,'success':' Data saved successfully !'})
        elif request.POST.get('cancel'):
            lid = request.POST.get('cancel')
            rem = request.POST.get(request.POST.get('cancel') + "_remarks")
            nof=float(Leave.objects.get(lid=lid).nof)
            ltype=Leave.objects.get(lid=lid).ltype
            eid=Leave.objects.get(lid=lid).eid_id
            r=-1.0
            if ltype == 0:
                r = (Employee.objects.get(e_id=eid).e_cl) + nof
                Employee.objects.filter(e_id=eid).update(e_cl=r)
            elif ltype == 1:
                r = (Employee.objects.get(e_id=eid).e_sl) + nof
                Employee.objects.filter(e_id=eid).update(e_sl=r)
            elif ltype == 2:
                r = (Employee.objects.get(e_id=eid).e_el) + nof
                Employee.objects.filter(e_id=eid).update(e_el=r)
            elif ltype == 3:
                r = (Employee.objects.get(e_id=eid).e_fl) + nof
                Employee.objects.filter(e_id=eid).update(e_fl=r)
            #print(str(lid)+" "+str(nof)+" "+str(ltype)+" "+str(r))
            Leave.objects.filter(lid=lid).update(status='R', updtby=logged_in,remarks=rem)
            lreqobj = Leave.objects.filter(queueid=logged_in).order_by('-reqdate')
            return render(request, 'ultimatix/employee/leaverequest.html', {'user': user, 'role': role, 'lreqobj': lreqobj,'success':' Data saved successfully !'})
        elif request.POST.get('fwd'):
            lreqobj = Leave.objects.filter(queueid=logged_in).order_by('-reqdate')
            lid=request.POST.get('fwd')
            fwdto=str(request.POST.get(request.POST.get('fwd')+"_fwdto"))
            if fwdto=="":
                return render(request, 'ultimatix/employee/leaverequest.html',
                              {'user': user, 'role': role, 'lreqobj': lreqobj,'error':' Please enter employee id'})
            else:
                ob=Employee.objects.filter(e_id=fwdto,e_status='Y')
                if not ob:
                    return render(request, 'ultimatix/employee/leaverequest.html',
                                  {'user': user, 'role': role, 'lreqobj': lreqobj, 'error': ' Please enter a valid employee id'})
                else:
                    ownid=Leave.objects.get(lid=lid).eid_id
                    if(str(fwdto)==str(ownid)):
                        return render(request, 'ultimatix/employee/leaverequest.html',
                                      {'user': user, 'role': role, 'lreqobj': lreqobj,'error': ' Invalid Employee !'})
                    else:
                        lid = request.POST.get('fwd')
                        rem=request.POST.get(request.POST.get('fwd') + "_remarks")
                        queueid=fwdto
                        Leave.objects.filter(lid=lid).update(queueid=queueid,updtby=logged_in,
                                                             remarks=rem)
                        return render(request, 'ultimatix/employee/leaverequest.html',
                                      {'user': user, 'role': role, 'lreqobj': lreqobj, 'success': ' Data saved successfully !'})
    else:
        lreqobj = Leave.objects.filter(queueid=logged_in).order_by('-reqdate')
        return render(request, 'ultimatix/employee/leaverequest.html', {'user': user,'role':role, 'lreqobj': lreqobj})


def empcalendar(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    if request.method=='POST':
        if request.POST.get('submit'):
            month=(request.POST.get('month_select'))
            m=month
            year =(request.POST.get('year_select'))
            ed = str(calendar.monthrange(int(year), int(month))[1])
            if len(month) == 1:
                month = '0' + month
            sdate = year + "-" + month + "-01"
            edate = year + "-" + month + "-" + ed

            cal = calendar.Calendar()
            hobj = Holidays.objects.filter(hdate__gte=sdate, hdate__lte=edate)
            holidays = []
            for o in hobj:
                holidays.append(o.hdate.day)
            tobj = Timesheet.objects.filter(eid=logged_in, tdate__gte=sdate, tdate__lte=edate)
            absent = []
            halfday = []
            for o in tobj:
                if o.tstate == 'A':
                    absent.append(o.tdate.day)
                elif o.tstate == 'H':
                    halfday.append(o.tdate.day)
            calendar_list = []
            for day in cal.itermonthdays(int(year), int(month)):
                t = {}
                t['date'] = day
                if day in holidays:
                    t['value'] = 0
                elif day in absent:
                    t['value'] = 1
                elif day in halfday:
                    t['value'] = 2
                else:
                    t['value'] = 3
                calendar_list.append(t)
            return render(request, 'ultimatix/employee/calendar.html',
                          {'user': user, 'role': role, 'calendar_list': calendar_list, 'hobj': hobj,'month':m,'year':year})
            #return HttpResponse(month+" "+year)
    else:
        date = datetime.now()
        month = str(date.month)
        m=month
        year = str(date.year)
        ed = str(calendar.monthrange(int(year), int(month))[1])
        if len(month) == 1:
            month = '0' + month
        sdate = year + "-" + month + "-01"
        edate = year + "-" + month + "-" + ed

        cal = calendar.Calendar()
        hobj=Holidays.objects.filter(hdate__gte=sdate, hdate__lte=edate)
        holidays=[]
        for o in hobj:
            holidays.append(o.hdate.day)
        tobj= Timesheet.objects.filter(eid=logged_in,tdate__gte=sdate, tdate__lte=edate)
        absent=[]
        halfday=[]
        for o in tobj:
            if o.tstate == 'A':
                absent.append(o.tdate.day)
            elif o.tstate == 'H':
                halfday.append(o.tdate.day)
        calendar_list=[]
        for day in cal.itermonthdays(int(year), int(month)):
            t={}
            t['date']=day
            if day in holidays:
                t['value']=0
            elif day in absent:
                t['value']=1
            elif day in halfday:
                t['value']=2
            else:
                t['value']=3
            calendar_list.append(t)
        #print(str(calendar_list))
        return render(request, 'ultimatix/employee/calendar.html', {'user': user, 'role': role,'calendar_list':calendar_list,'hobj':hobj,'month':m,'year':year})

def emppayslip(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    if request.method=='POST':
        adate=request.POST.get('sdate')
        r=Wage.objects.filter(eid_id=logged_in,sdate__lte=adate,
                              edate__gte=adate,status='Y')
        cursor = connection.cursor()
        if r:
            sql = '''
            select ue.e_id, (ue.e_fname ||' ' ||ifnull(ue.e_lname,'')) as Name, ud.d_desc, ues.es_ctc
            from ultimatix_employee ue
            left outer join ultimatix_employee_desig ued on ued.ed_eid_id = ue.e_id
            left outer join ultimatix_designation ud on ud.d_id = ued.ed_did_id
            left outer join ultimatix_employee_sal ues on ues.es_eid_id = ue.e_id
            where ued.ed_status = 'Y' and ues.es_status = 'Y' and ue.e_id = %s
            '''
            cursor.execute(sql,[logged_in])
            res=cursor.fetchone()
            t={}
            t['eid']=res[0]
            t['name']=res[1]
            t['desig']=res[2]
            t['ctc']=res[3]
            sql='''
            select bs,conv,hra,city,sundry,ptax,pf,esis,lop from ultimatix_wage where eid_id=%s and 
            status='Y' and sdate<=%s and edate>=%s
            '''
            cursor.execute(sql, [logged_in,adate,adate])
            res = cursor.fetchone()
            t['bs']=res[0]
            t['conv']=res[1]
            t['hra']=res[2]
            t['city']=res[3]
            t['sundry']=res[4]
            t['ptax']=res[5]
            t['pf']=res[6]
            t['esis']=res[7]
            t['lop']=res[8]
            t['earn']=float(res[0])+float(res[1])+float(res[2])+float(res[3])+float(res[4])
            t['ded']=float(res[5])+float(res[6])+float(res[7])
            t['fin']=round((float(res[0])+float(res[1])+float(res[2])+float(res[3])+float(res[4]))-float(res[8]),2)
            cursor.close()
            date=datetime.strptime(adate,'%Y-%m-%d')
            mnyr = date.strftime('%B') + ", " + date.strftime('%Y')
            t['mnyr']=mnyr
            pdf = render_to_pdf('pdf/payslip.html', t)
            return HttpResponse(pdf, content_type='application/pdf')
        else:
            return render(request, 'ultimatix/employee/payslip.html',
                  {'user': user, 'role': role,'error':' Payslip is not yet generated !'})
    else:
        return render(request, 'ultimatix/employee/payslip.html',
                  {'user': user, 'role': role})

def emptsupdt(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    if request.method=='POST':
        eid=request.POST.get('emp_select')
        tdate=request.POST.get('tdate')
        tstate=request.POST.get('attn_select')
        temp = Timesheet.objects.filter(eid=eid, tdate=tdate)
        if not temp:
            ob = Timesheet.objects.create(eid=Employee.objects.get(e_id=eid),tdate=tdate,tstate=tstate)
            ob.save()
        else:
            Timesheet.objects.filter(eid=Employee.objects.get(e_id=eid),tdate=tdate).update(tstate=tstate)

        pm = Project_members.objects.filter(eid_id=logged_in, pmstatus='Y')
        emp_list = []
        if pm:
            pid = Project_members.objects.get(eid_id=logged_in, pmstatus='Y').pid_id
            cursor = connection.cursor()
            sql = '''
            select upm.eid_id
            from ultimatix_project_members upm 
            where upm.pid_id=%s and upm.pmstatus='Y' and upm.eid_id not in ('1022890',%s) order by upm.eid_id            
            '''
            cursor.execute(sql, [pid, logged_in])
            res = cursor.fetchall()
            for r in res:
                emp_list.append(r[0])
        return render(request, 'ultimatix/employee/updttimesheet.html',
                      {'user': user, 'role': role, 'emp_list': emp_list,'success':' Successfully updated timesheet data, Please ask employee to verify !'})
        #return HttpResponse('success')
    else:
        pm=Project_members.objects.filter(eid_id=logged_in, pmstatus='Y')
        emp_list=[]
        if pm:
            pid = Project_members.objects.get(eid_id=logged_in, pmstatus='Y').pid_id
            cursor=connection.cursor()
            sql='''
            select upm.eid_id
            from ultimatix_project_members upm 
            where upm.pid_id=%s and upm.pmstatus='Y' and upm.eid_id not in ('1022890',%s) order by upm.eid_id            
            '''
            cursor.execute(sql,[pid,logged_in])
            res=cursor.fetchall()
            for r in res:
                emp_list.append(r[0])
            #emp_list=Project_members.objects.filter(pid_id=pid,pmstatus='Y')
        return render(request, 'ultimatix/employee/updttimesheet.html',{'user': user, 'role': role,'emp_list':emp_list})


def updttmsheet(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    #role=findrole(logged_in)
    if request.method=='POST':
        eid=request.POST.get('emp_select')
        tdate=request.POST.get('tdate')
        tstate=request.POST.get('attn_select')
        temp = Timesheet.objects.filter(eid=eid, tdate=tdate)
        if not temp:
            ob = Timesheet.objects.create(eid=Employee.objects.get(e_id=eid),tdate=tdate,tstate=tstate)
            ob.save()
        else:
            Timesheet.objects.filter(eid=Employee.objects.get(e_id=eid),tdate=tdate).update(tstate=tstate)
        emp_list = Employee.objects.filter(~Q(e_id='1022890'))
        return render(request, 'ultimatix/admin/updttimesheet.html',
                      {'user': user,'emp_list': emp_list,'success':' Successfully updated timesheet data'})
        #return HttpResponse('success')
    else:
        emp_list=Employee.objects.filter(~Q(e_id='1022890'))
        return render(request, 'ultimatix/admin/updttimesheet.html',
                      {'user': user,'emp_list': emp_list})

def empdtlupdt(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    if request.is_ajax() and request.GET:
        id = request.GET.get('id')
        cursor = connection.cursor()
        sql = """
        select e_id,e_fname,e_lname,e_sname,e_dob,e_bgroup,e_mstatus,e_nationality,e_qfn,e_disb,e_gender
        from ultimatix_employee
        where e_id=%s
        """
        cursor.execute(sql, [id])
        res = cursor.fetchall()
        res_list = []
        for r in res:
            t={}
            t['eid']=r[0]
            t['fname']=r[1]
            t['lname'] = r[2]
            t['sname'] = r[3]
            t['dob'] = str(r[4])
            t['bgroup'] = r[5]
            t['mstatus'] = r[6]
            t['nationality'] = r[7]
            t['qfn'] = r[8]
            t['disb'] = r[9]
            t['gender'] = r[10]
            res_list.append(t)
        j=json.dumps(res_list)
        cursor.close()
        #print(res_list)
        return JsonResponse(res_list, safe=False)
    elif request.method=='POST':
        eid = request.POST.get('emp_select')
        lname = request.POST.get('lname')
        sname = request.POST.get('surname')
        bgroup = request.POST.get('bgroup')
        mstatus = request.POST.get('mstatus')
        nationality = request.POST.get('nationality')
        qfn = request.POST.get('qualific')
        disb = request.POST.get('chk_disb')
        if not disb:
            disb = 'N'
        prdisb=Employee.objects.get(e_id=eid).e_disb
        if prdisb =='Y':
            disb='Y'

        Employee.objects.filter(e_id=eid).update(e_lname=lname, e_sname=sname, e_bgroup=bgroup,
                                                 e_mstatus=mstatus, e_disb=disb, e_qfn=qfn,e_nationality=nationality)

        #return HttpResponse('success')
        pm = Project_members.objects.filter(eid_id=logged_in, pmstatus='Y')
        emp_list=[]
        if pm:
            pid = Project_members.objects.get(eid_id=logged_in, pmstatus='Y').pid_id
            emp_list=findmyprojectmembers(pid,logged_in)
        return render(request,'ultimatix/employee/updtemp.html',{'user':user,'role':role,'emp_list':emp_list,'success':' Successfully updated employee details !'})
    else:
        pm = Project_members.objects.filter(eid_id=logged_in, pmstatus='Y')
        emp_list=[]
        if pm:
            pid = Project_members.objects.get(eid_id=logged_in, pmstatus='Y').pid_id
            emp_list=findmyprojectmembers(pid,logged_in)
        return render(request,'ultimatix/employee/updtemp.html',{'user':user,'role':role,'emp_list':emp_list})

def findmyprojectmembers(pid,logged_in):
    cursor = connection.cursor()
    sql = '''
    select upm.eid_id
    from ultimatix_project_members upm 
    where upm.pid_id=%s and upm.pmstatus='Y' and upm.eid_id not in ('1022890',%s) order by upm.eid_id            
    '''
    cursor.execute(sql, [pid, logged_in])
    res = cursor.fetchall()
    emp_list = []
    for r in res:
        emp_list.append(r[0])
    return emp_list

def admindashboard(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    e=Employee.objects.filter(~Q(e_id='1022890',e_status='Y')).count()
    p=Project.objects.filter(pstatus='A').count()
    #r=Project.objects.filter(pstatus='A').aggregate(Sum('pbudget'))
    r = Project.objects.aggregate(Sum('pbudget'))
    r1=round(float(r['pbudget__sum']/10000000),2)
    cursor = connection.cursor()
    sql="""
    select count(DISTINCT ue.e_id)
    from ultimatix_employee ue
    left outer join ultimatix_project_members upm on upm.eid_id=ue.e_id
    where upm.pmstatus='Y' and ue.e_id!='1022890'
    """
    cursor.execute(sql)
    res = cursor.fetchone()
    a=int(res[0])
    print(str(e)+" "+str(a))
    u=e-a
    sql='''
    select ed_did_id,ud.d_cd,count(*) as tot_count from ultimatix_employee_desig ued 
    left outer join ultimatix_designation ud on ud.d_id = ued.ed_did_id
    where ud.d_cd!='SYSADMIN' and ued.ed_status='Y'
    group by ued.ed_did_id order by ed_did_id
    '''
    cursor.execute(sql)
    res = cursor.fetchall()

    '''label=""
    data=""
    for r in res:
        label=label+'"'+str(r[1])+'"'+","
        data=data+'"'+str(r[2])+'"'+","
    label = label[:-1]
    data=data[:-1]
    print(label)'''

    for r in res:
        if str(r[1])=='ASE-T':
            ase_t=int(r[2])
        elif str(r[1])=='ASE':
            ase=int(r[2])
        elif str(r[1])=='SE':
            se=int(r[2])
        elif str(r[1])=='IT-A':
            it_a=int(r[2])
        elif str(r[1])=='AST-C':
            ast_c=int(r[2])
        elif str(r[1])=='ASC-C':
            asc_c=int(r[2])
        elif str(r[1])=='CT':
            ct=int(r[2])
        elif str(r[1])=='HR':
            hr=int(r[2])

    pact=Project.objects.filter(pstatus='A').count()
    pcomp=Project.objects.filter(pstatus='C').count()
    return render(request,'ultimatix/admin/dashboard.html',{'user':user,'e':e,'p':p,
                        'r1':r1,'u':u,'a':a,'ase_t':ase_t,'ase':ase,'se':se,'it_a':it_a,
                        'ast_c':ast_c,'asc_c':asc_c,'ct':ct,'hr':hr,'pact':pact,'pcomp':pcomp})


def salary_details(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    if request.method=='POST':
        month = (request.POST.get('month_select'))
        m = month
        year = (request.POST.get('year_select'))
        ed = str(calendar.monthrange(int(year), int(month))[1])
        if len(month) == 1:
            month = '0' + month
        sdate = year + "-" + month + "-01"
        edate = year + "-" + month + "-" + ed

        cursor = connection.cursor()
        sql='''
        select ue.e_id, (ue.e_fname ||' ' ||ue.e_lname) as Name, ud.d_desc,uw.ctc,
        uw.ctc_pm,
        uw.bs,uw.conv,uw.hra,uw.city,uw.sundry,
        uw.ptax,uw.pf,uw.esis,uw.lop,
        uw.sdate,uw.edate
        from ultimatix_employee ue
        left outer join ultimatix_employee_desig ued on ued.ed_eid_id = ue.e_id
        left outer join ultimatix_designation ud on ud.d_id = ued.ed_did_id
        left outer join ultimatix_wage uw on uw.eid_id=ue.e_id
        where ued.ed_status = 'Y' and ud.d_cd!='SYSADMIN' and uw.sdate>=%s and uw.edate<=%s        
        '''
        cursor.execute(sql,[sdate,edate])
        res = cursor.fetchall()
        res_list=[]
        s=0.0
        print(str(res))
        for r in res:
            t={}
            t['eid']=r[0]
            t['name']=r[1]
            t['desig']=r[2]
            t['ctc']=float(r[3])
            t['ctc_pm']=float(r[4])
            t['bs']=float(r[5])
            t['conv']=float(r[6])
            t['hra']=float(r[7])
            t['city']=float(r[8])
            t['sundry']=float(r[9])
            t['ptax']=float(r[10])
            t['pf']=float(r[11])
            t['esis']=float(r[12])
            t['lop']=float(r[13])
            #t['earn']=float(res[5])+float(res[6])+float(res[7])+float(res[8])+float(res[9])
            t['earn']=round((t['bs']+t['conv']+t['hra']+t['city']+t['sundry']),2)
            #t['ded']=float(res[10])+float(res[11])+float(res[12])
            t['ded']=round((t['ptax']+t['pf']+t['esis']),2)
            #t['fin']=round((float(res[5])+float(res[6])+float(res[7])+float(res[8])+float(res[9]))-float(res[13]),2)
            t['fin']=round((t['earn']-t['lop']),2)
            #t['earn']=float(t['bs'])
            #print(t['earn'])
            #s=s+round((float(res[5])+float(res[6])+float(res[7])+float(res[8])+float(res[9]))-float(res[13]),2)
            res_list.append(t)
            s=s+(t['earn']-t['lop'])
        return render(request, 'ultimatix/admin/salary_details.html', {'user': user,'res_list':res_list,'s':s,'month':m,'year':year})
    else:
        return render(request, 'ultimatix/admin/salary_details.html',{'user':user})


def empprofile(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    if request.method=='POST':
        paddr1 = request.POST.get('paddr1')
        paddr2 = request.POST.get('paddr2')
        paddr3 = request.POST.get('paddr3')
        caddr1 = request.POST.get('caddr1')
        caddr2 = request.POST.get('caddr2')
        caddr3 = request.POST.get('caddr3')
        email = request.POST.get('email')
        mnumber = request.POST.get('mnumber')
        pwd=request.POST.get('psw1')
        #img=request.FILES['myfile']
        #print(img)
        if pwd=="":
            pwd=user.e_pwd
        #print(pwd)
        '''if img:
            Employee.objects.filter(e_id=logged_in).update(e_paddr1=paddr1,
                                                 e_paddr2=paddr2,e_paddr3=paddr3,e_caddr1=caddr1,
                                                 e_caddr2=caddr2,e_caddr3=caddr3,e_email=email,
                                               e_mnumber=mnumber,e_pwd=pwd,e_img=img)
        else:'''
        Employee.objects.filter(e_id=logged_in).update(e_paddr1=paddr1,
                                             e_paddr2=paddr2,e_paddr3=paddr3,e_caddr1=caddr1,
                                             e_caddr2=caddr2,e_caddr3=caddr3,e_email=email,
                                           e_mnumber=mnumber,e_pwd=pwd)
        return redirect('ultimatix:home_view')
        #return HttpResponse('success')
    else:
        return render(request,'ultimatix/employee/profile.html',{'user':user,'role':role})


def empappraisal(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    if request.method=='POST':
        date = datetime.now().date()
        s = str(date.year - 1) + "-01-01"
        ed = calendar.monthrange(date.year - 1, 12)[1]
        e = str(date.year - 1) + "-12-" + str(ed)
        m = str(date.month)
        if len(m) == 1:
            m = '0' + str(m)
        date1 = str(date.year - 1) + "-" + m + "-" + str(date.day)
        date2=str(Employee.objects.get(e_id=logged_in).e_doj)
        print(date2)
        print(date)
        d2=datetime.strptime(str(date2), '%Y-%m-%d')
        d1=datetime.strptime(str(date), '%Y-%m-%d')
        print(d2)
        print(d1)
        print(str((d1-d2).days))
        days1=int((d1-d2).days)
        '''if days1<180:
            aobj = Appraisal.objects.filter(eid=logged_in)
            return render(request, 'ultimatix/employee/appraisal.html', {'user': user, 'role': role,'aobj':aobj,'error':' You cant initiate appraisal now !'})'''
        ob=Appraisal.objects.filter(eid=logged_in,asdate__lte=date1,aedate__gte=date1)
        if not ob:
            appr=findqueueid(logged_in)
            Appraisal.objects.create(eid=Employee.objects.get(e_id=logged_in),
                                     appraiseeid=appr,asdate=s,aedate=e,status='P')
            aobj=Appraisal.objects.filter(eid=logged_in)
            return render(request, 'ultimatix/employee/appraisal.html', {'user': user, 'role': role,'aobj':aobj,'success':'Initiated successfully !'})
        else:
            aobj = Appraisal.objects.filter(eid=logged_in)
            return render(request, 'ultimatix/employee/appraisal.html', {'user': user, 'role': role,'aobj':aobj,'error':'Appraisal has been initiated already for the same period ! !'})
    else:
        aobj = Appraisal.objects.filter(eid=logged_in)
        return render(request,'ultimatix/employee/appraisal.html',{'user':user,'role':role,'aobj':aobj})


def empappraisalreq(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    if request.method=='POST':
        if request.POST.get('submit'):
            aid = request.POST.get('submit')
            rating = float(request.POST.get(request.POST.get('submit') + "_rating"))
            comments= request.POST.get(request.POST.get('submit') + "_comments")
            eid=Appraisal.objects.get(aid=aid).eid_id
            prev_sal=float(Employee_sal.objects.get(es_eid=eid,es_status='Y').es_ctc)
            if rating==10:
                inr=12.0
            elif rating>=8:
                inr=10.0;
            elif rating>=6:
                inr=7.0
            elif rating>=4:
                inr=5.0
            elif rating>=2:
                inr=3.0
            else:
                inr=1.0
            #print(inr)
            new_sal=float(prev_sal)+(float(prev_sal)*inr)/100
            Employee_sal.objects.filter(es_eid=eid,es_status='Y').update(es_ctc=new_sal)
            Appraisal.objects.filter(aid=aid).update(rating=rating,comments=comments,status='C',updtby=logged_in)
            aobj = Appraisal.objects.filter(appraiseeid=logged_in, status='P')
            return render(request, 'ultimatix/employee/appraisalreq.html', {'user': user, 'role': role, 'aobj': aobj})
            #return HttpResponse(str(new_sal))
        elif request.POST.get('fwd'):
            aobj = Appraisal.objects.filter(appraiseeid=logged_in,status='P')
            aid=request.POST.get('fwd')
            fwdto=str(request.POST.get(request.POST.get('fwd')+"_fwdto"))
            if fwdto=="":
                return render(request, 'ultimatix/employee/appraisalreq.html',
                              {'user': user, 'role': role,'aobj': aobj,'error':' Please enter employee id'})
            else:
                ob=Employee.objects.filter(e_id=fwdto,e_status='Y')
                if not ob:
                    return render(request, 'ultimatix/employee/appraisalreq.html',
                                  {'user': user, 'role': role, 'aobj': aobj, 'error': ' Please enter a valid employee id'})
                else:
                    ownid=Appraisal.objects.get(aid=aid).eid_id
                    #print(str(ownid)+" "+str(fwdto))
                    if(str(fwdto)==str(ownid)):
                        return render(request, 'ultimatix/employee/appraisalreq.html',
                                      {'user': user, 'role': role, 'aobj': aobj,'error': ' Invalid Employee !'})
                    else:
                        aid = request.POST.get('fwd')
                        rating=request.POST.get(request.POST.get('fwd') + "_rating")
                        #print(float(rating))
                        comments=request.POST.get(request.POST.get('fwd') + "_comments")
                        queueid=fwdto
                        Appraisal.objects.filter(aid=aid).update(appraiseeid=queueid,updtby=logged_in,
                                                             rating=float(rating),comments=comments)
                        aobj = Appraisal.objects.filter(appraiseeid=logged_in,status='P')
                        return render(request, 'ultimatix/employee/leaverequest.html',
                                      {'user': user, 'role': role, 'aobj': aobj, 'success': ' Data saved successfully !'})
            #return HttpResponse(str(aid) + str(fwdto))
    else:
        aobj = Appraisal.objects.filter(appraiseeid=logged_in,status='P')
        return render(request, 'ultimatix/employee/appraisalreq.html', {'user': user, 'role': role, 'aobj': aobj})

def adminappraisalreq(request):
    logged_in = request.session['eid']
    user = Employee.objects.get(e_id=logged_in)
    role=findrole(logged_in)
    if request.method=='POST':
        if request.POST.get('submit'):
            aid = request.POST.get('submit')
            rating = float(request.POST.get(request.POST.get('submit') + "_rating"))
            comments= request.POST.get(request.POST.get('submit') + "_comments")
            eid=Appraisal.objects.get(aid=aid).eid_id
            prev_sal=float(Employee_sal.objects.get(es_eid=eid,es_status='Y').es_ctc)
            if rating==10:
                inr=12.0
            elif rating>=8:
                inr=10.0;
            elif rating>=6:
                inr=7.0
            elif rating>=4:
                inr=5.0
            elif rating>=2:
                inr=3.0
            else:
                inr=1.0
            #print(inr)
            new_sal=float(prev_sal)+(float(prev_sal)*inr)/100
            Employee_sal.objects.filter(es_eid=eid,es_status='Y').update(es_ctc=new_sal)
            Appraisal.objects.filter(aid=aid).update(rating=rating,comments=comments,status='C',updtby=logged_in)
            aobj = Appraisal.objects.filter(appraiseeid=logged_in, status='P')
            return render(request, 'ultimatix/admin/appraisalreq.html', {'user': user, 'role': role, 'aobj': aobj})
    else:
        aobj = Appraisal.objects.filter(appraiseeid=logged_in, status='P')
        return render(request, 'ultimatix/admin/appraisalreq.html', {'user': user, 'role': role, 'aobj': aobj})